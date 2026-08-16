#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Controlelijst: RGD-projecten met een adres (locatie) dat nog niet op adres/straat-niveau
gekoppeld is (alleen plaats, of helemaal geen coord). Voor handmatige correctie.
Schrijft rgd_controlelijst.xlsx met lege kolommen 'juist adres / lat / lon'."""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORK = os.path.dirname(os.path.abspath(__file__))
A = json.load(open(f"{WORK}/rgd_all.json", encoding="utf-8"))
NA = "https://www.nationaalarchief.nl/onderzoeken/archief/4.RGD/invnr/{}/file/"

def loc(p): return (p.get("locatie") or "").strip()
# adres bekend, maar geocode alleen op plaats-niveau of ontbreekt
rows = [p for p in A if loc(p) and (p.get("prec") in ("plaats", None))]
rows.sort(key=lambda p: (p.get("invnr") or "", p.get("stad") or ""))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "te koppelen"
head = ["inv.nr", "gebouw", "adres (locatie)", "stad", "huidige precisie",
        "huidige lat", "huidige lon", "NA-link", "→ juist adres", "→ juiste lat", "→ juiste lon"]
ws.append(head)
hf = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="0E7490")
for c in ws[1]:
    c.font = hf; c.fill = fill; c.alignment = Alignment(vertical="center")
warn = PatternFill("solid", fgColor="FDE68A")
for p in rows:
    inv = p.get("invnr") or (p.get("uid") or "").split("-")[0]
    ws.append([inv, p.get("gebouw") or p.get("titel", ""), loc(p), p.get("stad", ""),
               p.get("prec") or "geen", p.get("lat"), p.get("lon"),
               NA.format(inv), "", "", ""])
    if p.get("prec") in (None,):                        # geen coord: geel markeren
        ws.cell(ws.max_row, 5).fill = warn
widths = [9, 34, 30, 18, 14, 11, 11, 60, 30, 12, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
out = "/Users/alan/Downloads/rgd_controlelijst.xlsx"
wb.save(out)
print(f"geschreven: {out}  ({len(rows)} rijen; waarvan {sum(1 for p in rows if p.get('prec') is None)} zonder coord)")
