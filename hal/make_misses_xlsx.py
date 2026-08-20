# -*- coding: utf-8 -*-
"""Voeg alle misses/<periode>.json samen tot 1 Excel voor handmatige 2e check."""
import json,collections
import openpyxl
from openpyxl.styles import Font, PatternFill
d=json.load(open('misses_current.json')); per=d['periode']
agg=collections.Counter(); periods=collections.defaultdict(set)
for naar,c in d['misses']:
    agg[naar]+=c; periods[naar].add(per)
wb=openpyxl.Workbook(); ws=wb.active; ws.title='niet gematcht'
hdr=['Naar (ruwe waarde)','Aantal (totaal)','Periode(s)','Mogelijke plaats','Land','Lat','Lon','Opmerking']
ws.append(hdr)
for i,c in enumerate(hdr,1):
    cell=ws.cell(1,i); cell.font=Font(bold=True); cell.fill=PatternFill('solid',fgColor='0E7490'); cell.font=Font(bold=True,color='FFFFFF')
for naar,c in agg.most_common():
    ws.append([naar,c,', '.join(sorted(periods[naar])),'','','','',''])
ws.freeze_panes='A2'
widths=[34,14,16,24,10,10,10,30]
for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
ws.auto_filter.ref=f'A1:H{ws.max_row}'
out='HAL_niet_gematchte_bestemmingen.xlsx'
wb.save(out)
print(f"{out}: {len(agg)} unieke waarden, {sum(agg.values())} vermeldingen totaal")
