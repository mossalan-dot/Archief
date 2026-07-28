#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Detecteer makkelijk vindbare fouten in de dossierbeschrijvingen (2.19.364)
en schrijf ze naar Carnegie-mogelijke-fouten.xlsx (inv.nr. voorop, NA-link per rij).

Hoge-precisie-detectoren (geen uitputtende controle):
  1. plaatsnaam-spelfouten  — bekende bron/OCR-fouten uit de geocode-aliassen
  2. plaatsnaam-onzeker     — plaatsen die we niet betrouwbaar konden thuisbrengen
  3. dubbele woorden        — "te te", "een een", ...
  4. gesplitste woorden     — OCR-splitsing met een echte spatie ("afgewe zen")
"""
import json, os, re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORK = os.path.dirname(os.path.abspath(__file__))
recs = json.load(open(f"{WORK}/dossiers_raw.json", encoding="utf-8"))
NA = "https://www.nationaalarchief.nl/onderzoeken/archief/2.19.364/invnr/"

# echte spelfouten in de bron (niet: verouderde gemeentenamen, die zijn correct)
SPELFOUT = {
    "Snek": "Sneek", "Ursum": "Ursem", "Wansum": "Wanssum", "Wychen": "Wijchen",
    "Galadammen": "Galamadammen", "'s-Gravenhagen": "'s-Gravenhage", "'s-Gravenhag": "'s-Gravenhage",
    "Herenveen": "Heerenveen", "Veendendaal": "Veenendaal", "Eindhoen": "Eindhoven", "Yrseke": "Yerseke",
    "Warfum": "Warffum", "Standaarbuiten": "Standdaarbuiten", "Rijsburg": "Rijnsburg",
    "Etten-Leuk": "Etten-Leur", "St Michielgestel": "Sint-Michielsgestel", "Nieuwehoorn": "Nieuwenhoorn",
    "Balkburg": "Balkbrug", "Veenwoude": "Veenwouden", "Bennebroer": "Bennebroek", "Nijholtpa": "Nijeholtpade",
    "Jutpaas": "Jutphaas", "Bassingerhorn": "Barsingerhorn", "Kootsertille": "Kootstertille",
    "Bergummerdam": "Bergumerdam", "Wijntjeterp": "Wijnjeterp", "Emmercompascuum": "Emmer-Compascuum",
    "Gasselte-Nijveen": "Gasselternijveen", "Harbrinkbroek": "Harbrinkhoek", "Nieuwelande": "Nieuwlande",
    "Rinsumageest": "Rinsumageast", "Terheijde aan Zee": "Ter Heijde", "Dirkland": "Dirksland",
    "Wervershoef": "Wervershoof", "Otterleek": "Oterleek", "Wiedum": "Weidum",
    "Ouderwerk aan den IJssel": "Ouderkerk aan den IJssel", "Breukelerveen": "Breukeleveen",
    "Achtkaspelen": "Achtkarspelen", "Datumadeel": "Dantumadeel", "Norresunby": "Nørresundby",
    "Tandjong Priok": "Tanjung Priok", "Tjipatoedjah": "Cipatujah", "Flambourough Head": "Flamborough Head",
    "Haiming Titol": "Haiming (Tirol)", "Nieuw-Ginniken": "Ginneken",
}
ONZEKER = {
    "Sluis-Buitenlust": "niet thuisgebracht", "Altenhuntroff": "niet thuisgebracht (Duitsland?)",
    "St. Lucia": "niet thuisgebracht (niet in NL)", "Lijtshuizen": "gehucht bij IJlst (Wymbritseradeel)",
}
SPLITS = [("afgewe zen", "afgewezen"), ("onderste un", "ondersteun"), ("gratifi catie", "gratificatie"),
          ("kin deren", "kinderen"), ("water van de de", "water van de")]

def snip(d): return (d[:120] + "…") if len(d) > 120 else d

rows = []
for x in recs:
    nr, d, y = x["nr"], x["desc"], x.get("year_from") or ""
    for p in x["places"]:
        if p in SPELFOUT:  rows.append([nr, "plaatsnaam (spelling)", p, SPELFOUT[p], y, snip(d)])
        elif p in ONZEKER: rows.append([nr, "plaatsnaam (onzeker)", p, ONZEKER[p], y, snip(d)])
    m = re.search(r"\b(\w{2,})\s+\1\b", d, re.I)
    if m and m.group(1).lower() not in ("de", "het", "en"):
        rows.append([nr, "dubbel woord", m.group(0), m.group(1), y, snip(d)])
    for bad, good in SPLITS:
        if re.search(r"\b" + bad.replace(" ", r"\s+") + r"\b", d, re.I):
            rows.append([nr, "gesplitst woord", bad, good, y, snip(d)]); break

CATORD = {"plaatsnaam (spelling)": 0, "plaatsnaam (onzeker)": 1, "gesplitst woord": 2, "dubbel woord": 3}
rows.sort(key=lambda r: (CATORD.get(r[1], 9), r[0]))
cnt = Counter(r[1] for r in rows)

# ---- Excel ----
FONT = "Arial"; ink = Font(name=FONT, size=10)
CATFILL = {"plaatsnaam (spelling)": "FCE4D6", "plaatsnaam (onzeker)": "FFF2CC",
           "gesplitst woord": "E2EFDA", "dubbel woord": "DDEBF7"}
wb = Workbook(); ws = wb.active; ws.title = "Mogelijke fouten"
ws.merge_cells("A1:G1")
ws["A1"] = "Carnegie Heldenfonds — mogelijke fouten in de dossierbeschrijvingen (NA 2.19.364)"
ws["A1"].font = Font(name=FONT, size=13, bold=True); ws["A1"].alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 22
ws.merge_cells("A2:G2")
ws["A2"] = (f"Automatisch gedetecteerd ({len(rows)} items): {cnt['plaatsnaam (spelling)']} plaatsnaam-spelfouten, "
            f"{cnt['plaatsnaam (onzeker)']} onzekere plaatsen, {cnt['dubbel woord']} dubbele woorden, "
            f"{cnt['gesplitst woord']} gesplitste woorden. Kolom 'Waarschijnlijk correct' is een suggestie — "
            "controleer bij twijfel. Klik het inv.nr. om het dossier bij het NA te openen.")
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="555555")
ws["A2"].alignment = Alignment(vertical="center", wrap_text=True); ws.row_dimensions[2].height = 42

headers = ["Inv.nr.", "Categorie", "Gevonden in beschrijving", "Waarschijnlijk correct / opmerking",
           "Jaar", "Beschrijving (fragment)", "NA-dossier"]
hr = 4; hfill = PatternFill("solid", fgColor="1F4E79")
for j, h in enumerate(headers, 1):
    c = ws.cell(hr, j, h); c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
    c.fill = hfill; c.alignment = Alignment(vertical="center", wrap_text=True)
ws.row_dimensions[hr].height = 28

border = Border(bottom=Side(style="thin", color="D9D9D9"))
r = hr + 1
for nr, cat, found, corr, year, sn in rows:
    a = ws.cell(r, 1, nr); a.hyperlink = NA + str(nr)
    a.font = Font(name=FONT, size=10, bold=True, color="1155CC", underline="single")
    b = ws.cell(r, 2, cat); b.font = ink; b.fill = PatternFill("solid", fgColor=CATFILL.get(cat, "FFFFFF"))
    ws.cell(r, 3, found).font = Font(name=FONT, size=10, bold=True)
    ws.cell(r, 4, corr).font = ink
    ws.cell(r, 5, year).font = ink; ws.cell(r, 5).alignment = Alignment(horizontal="left")
    ws.cell(r, 6, sn).font = ink
    lk = ws.cell(r, 7, "openen ↗"); lk.hyperlink = NA + str(nr)
    lk.font = Font(name=FONT, size=10, color="1155CC", underline="single")
    for j in range(1, 8): ws.cell(r, j).border = border
    r += 1

for j, w in enumerate([9, 22, 26, 32, 7, 72, 12], 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A5"; ws.auto_filter.ref = f"A{hr}:G{r-1}"; ws.sheet_view.showGridLines = False
wb.save(f"{WORK}/Carnegie-mogelijke-fouten.xlsx")
print(f"geschreven: Carnegie-mogelijke-fouten.xlsx | {len(rows)} items | {dict(cnt)}")
